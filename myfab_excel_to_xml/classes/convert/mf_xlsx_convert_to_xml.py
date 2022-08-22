# -*- coding: utf-8 -*-
from openerp import models, api, fields, _
from openerp.exceptions import ValidationError

import base64
from StringIO import StringIO

# some help https://www.pythonexcel.com/openpyxl.php
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# some help https://docs.python.org/2.7/library/xml.etree.elementtree.html
import xml
import xml.etree.ElementTree as ET

class mf_xlsx_convert_to_xml(models.Model):
    _name = 'mf.xlsx.convert.xml'
        
    #===========================================================================
    # COLUMNS
    #===========================================================================
    name = fields.Char(required=True)
    configuration_id = fields.Many2one('mf.xlsx.configuration', string='Configuration', ondelete='restrict')
    xlsx_file = fields.Binary(string="XLSX file to convert", required=True)
    xlsx_file_name =  fields.Char()
    xml_file = fields.Binary(string="XML file converted", readonly=True)
    xml_file_name = fields.Char()
    execution_message = fields.Char(readonly=True)

    #===========================================================================
    # METHODS
    #===========================================================================
    @api.one
    def mf_convert(self):
        if self._is_parameters_valide():
            (self.execution_message,self.xml_file) = self._mf_convert_XLSX_to_XML(self.xlsx_file, self.configuration_id)

    #===========================================================================
    # XLSX TO XML METHODS
    #===========================================================================
    def _mf_convert_XLSX_to_XML(self, xlsx_file, configuration_id):
        conversion_message = []
        xml_file = StringIO('')

        # Load XLSX file
        try:
            xlsx_document = load_workbook(filename=StringIO(base64.decodestring(xlsx_file)), data_only=True)
        except:
            xlsx_document = False

        # Convert XLSX to XML
        if not xlsx_document:
            conversion_message.append(_('Error on loading XLSX file'))
        else:
            ET_root = ET.Element(configuration_id.root_beacon)
            for sheet_rc in configuration_id.sheet_ids:
                if not self._mf_convert_XLSX_sheet_to_XML(ET_root, xlsx_document, sheet_rc):
                    conversion_message.append(_('No XLSX sheet "%s"')%(sheet_rc.sheet_name_or_index))

        # Write XML file
        if len(conversion_message) == 0:
            #print(ET.dump(ET_root)) # Activate it only for test 
            try:
                xml_file_content = ET.tostring(ET_root, encoding="UTF-8", method="xml")     
                #print(xml_file_content)
                xml_file = StringIO(xml_file_content)       
            except:
                conversion_message.append(_('Error on writing XML file'))

        # Manage conversion messages
        if len(conversion_message) == 0:
            conversion_message = _('Conversion finished with success')
        else:
            conversion_message = '\n'.join(conversion_message)

        return (conversion_message,xml_file)

    def _mf_convert_XLSX_sheet_to_XML(self, ET_root, xlsx_document, sheet_id):
        xlsx_sheet = self._mf_set_XLSX_active_sheet(xlsx_document, sheet_id.sheet_name_or_index)
        if not xlsx_sheet: return False

        # Copy the XML element to not change the working cursor in main element
        # but it assign a reference, so at end the main element has all sub elements added
        xlsx_rows = self._mf_get_XLSX_rows(xlsx_sheet, sheet_id)
        if len(xlsx_rows) > 0:
            ET_sheet = self._mf_add_XML_last_sub_element(ET_root, sheet_id.beacon_for_sheet, False) if sheet_id.beacon_for_sheet else ET_root
        
        # store the treated xlsx rows in child treatment, to not do them again as parent rows
        xlsx_rows_in_xml = []
        for xlsx_row_index in range(len(xlsx_rows)):
            self._mf_add_XLSX_row_to_XML(ET_sheet, xlsx_sheet, xlsx_rows, xlsx_row_index, 
                                        sheet_id.beacon_per_row, sheet_id.field_ids, sheet_id.level_field_id, 
                                        xlsx_rows_in_xml)

        return True        

    def _mf_add_XLSX_row_to_XML(self, ET_sheet, xlsx_sheet, xlsx_rows, xlsx_row_index, row_beacon, field_ids, level_id, xlsx_rows_in_xml):       
        if xlsx_row_index < len(xlsx_rows) and xlsx_row_index not in xlsx_rows_in_xml:
            xlsx_row = xlsx_rows[xlsx_row_index]
            
            # Add only here the row index to the list, because we write data in XML
            xlsx_rows_in_xml.append(xlsx_row_index)
            # Add beacon with all fields value
            ET_row = self._mf_add_XML_last_sub_element(ET_sheet, row_beacon, True)
            for field_rc in field_ids:
                self._mf_set_XML_field_value(ET_row, xlsx_sheet, xlsx_row, field_rc)

            # After adding the current row we check the children
            if level_id:
                self._mf_convert_children_rows_to_xml(ET_row, xlsx_sheet, xlsx_rows, xlsx_row_index, 
                                                    row_beacon, field_ids, level_id, 
                                                    xlsx_rows_in_xml)
        
    def _mf_set_XML_field_value(self, ET_row, xlsx_sheet, xlsx_row, field_id):
        old_Value = ''
        new_value = self._mf_get_cell_value(xlsx_sheet, xlsx_row, field_id)
        if new_value:
            ET_field = self._mf_add_XML_last_sub_element(ET_row, field_id.beacon, False)
            if field_id.attribute:
                if field_id.is_merge: old_Value = ET_field.get(field_id.attribute, '')
                ET_field.set(field_id.attribute, old_Value + new_value)
            else:
                if field_id.is_merge: old_Value = ET_field.text
                ET_field.text = old_Value + new_value
    
    def _mf_convert_children_rows_to_xml(self, ET_row, xlsx_sheet, xlsx_rows, xlsx_row_index, row_beacon, field_ids, level_id, xlsx_rows_in_xml):
        if xlsx_row_index < len(xlsx_rows):
            xlsx_row = xlsx_rows[xlsx_row_index]

            if level_id:
                current_level_value = self._mf_get_XLSX_cell_value(xlsx_sheet, xlsx_row, level_id.column)
                # We check row per row if it is a child
                child_xlsx_row_index = xlsx_row_index + 1
                while self._mf_is_row_child(xlsx_sheet, xlsx_rows, child_xlsx_row_index, level_id):
                    if child_xlsx_row_index not in xlsx_rows_in_xml and self._mf_is_direct_child(xlsx_sheet, xlsx_rows, child_xlsx_row_index, level_id, current_level_value):
                        ET_Child = self._mf_add_XML_last_sub_element(ET_row, level_id.beacon_per_level, False)
                        self._mf_add_XLSX_row_to_XML(ET_Child, xlsx_sheet, xlsx_rows, child_xlsx_row_index, 
                                                    row_beacon, field_ids, level_id, 
                                                    xlsx_rows_in_xml) 
                    child_xlsx_row_index += 1

    def _mf_get_cell_value(self, xlsx_sheet, xlsx_row, field_id):
        if field_id.writing_mode == 'column_value':
            cell_value = self._mf_get_XLSX_cell_value(xlsx_sheet, xlsx_row, field_id.column)
        elif field_id.writing_mode == 'constant_value' and field_id.fixed_value:
            cell_value = field_id.fixed_value
        else: 
            cell_value = False

        return cell_value

    #===========================================================================
    # CHECK METHODS
    #===========================================================================
    def _is_parameters_valide(self):
        if not self.xlsx_file : raise ValidationError(_('No XLSX file to convert.'))
        if len(self.xlsx_file_name) <= 5 or self.xlsx_file_name[-5:].upper() != '.XLSX' : raise ValidationError(_('Conversion only works with XLSX files'))
        if not self.configuration_id: raise ValidationError(_('Choose a configuration.'))
        return True

    def _mf_is_row_child(self, xlsx_sheet, xlsx_rows, child_xlsx_row_index, level_id):
        if child_xlsx_row_index < len(xlsx_rows):
            xlsx_row = xlsx_rows[child_xlsx_row_index]

            if level_id.is_numerical_level:
                level_value = self._mf_get_XLSX_cell_value(xlsx_sheet, xlsx_row, level_id.column)
                return (len(level_value) > 1)
            else:   
                return self._mf_get_XLSX_cell_value(xlsx_sheet, xlsx_row, level_id.parent_reference_column)

    def _mf_is_direct_child(self, xlsx_sheet, xlsx_rows, child_xlsx_row_index, level_id, level_parent_value):
        if child_xlsx_row_index < len(xlsx_rows):
            xlsx_row = xlsx_rows[child_xlsx_row_index]

            if level_id.is_numerical_level:
                level_child_value = self._mf_get_XLSX_cell_value(xlsx_sheet, xlsx_row, level_id.column)
                return self._mf_is_direct_numerical_child(level_parent_value, level_child_value, level_id.level_separator)
            else:
                parent_value = self._mf_get_XLSX_cell_value(xlsx_sheet, xlsx_row, level_id.parent_reference_column)
                if level_parent_value == parent_value: 
                    return True
        return False
    
    def _mf_is_direct_numerical_child(self, parent_level_value, current_level_value, level_separator):
        parent_level_values = parent_level_value.split(level_separator)
        current_level_values = current_level_value.split(level_separator)

        if len(parent_level_values)+1 == len(current_level_values):
            is_same_level_as_parent = True
            for i in range(len(parent_level_values)):
                if parent_level_values[i] != current_level_values[i]:
                    is_same_level_as_parent = False
            return is_same_level_as_parent
        return False

    #===========================================================================
    # XML METHODS
    #===========================================================================
    def _mf_add_XML_last_sub_element(self, ET_root, beacon_str, add_last_beacon_as_list):
        ET_Temp = ET_root
        
        if beacon_str:
            beacon_list = beacon_str.split('/')

            cpt = 0
            # Check if each beacon is already existing, else add it
            for searched_beacon in beacon_list:
                cpt +=1
                is_beacon_present = False
                if add_last_beacon_as_list == False or cpt < len(beacon_list):
                    for child in ET_root:
                        if child.tag == searched_beacon : 
                            is_beacon_present = True
                            ET_Temp = child
                if is_beacon_present == False:
                    ET_Temp = ET.SubElement(ET_root, searched_beacon)
                ET_root = ET_Temp

        return ET_root

    #===========================================================================
    # XLSX METHODS
    #===========================================================================
    def _mf_set_XLSX_active_sheet(self, xlsx_document, sheet_name_or_index):
        xlsx_sheet = False
        try:
            if sheet_name_or_index.isnumeric():
                xlsx_document.active = int(sheet_name_or_index) - 1 #index starts at 0 but user often gives index started by 1
                xlsx_sheet = xlsx_document.active
            else:
                xlsx_sheet = xlsx_document[sheet_name_or_index]
        except:
            return False
        
        return xlsx_sheet

    def _mf_get_XLSX_rows(self, xlsx_sheet, sheet_id):
        ending_line = sheet_id.ending_line if sheet_id.ending_line else xlsx_sheet.max_row
        xlsx_rows = []
        for xlsx_row in range(sheet_id.starting_line, ending_line +1):
            xlsx_rows.append(xlsx_row)
        return xlsx_rows

    def _mf_get_XLSX_cell_value(self, xlsx_sheet, xlsx_row, column):
        cell_value = xlsx_sheet['%s%s'%(column, xlsx_row)].value
        if cell_value == None: return False
    
        #Convert value readed in cell to be used by the rest of the functions
        """"
        try:
            if ',' in cell_value:
                cell_value = float(cell_value)
            else:
                cell_value = int(cell_value)
        except:
                cell_value = unicode(cell_value)
        """
        cell_value = unicode(cell_value)

        if cell_value == '': return False

        cell_value = cell_value.strip()
        return cell_value

